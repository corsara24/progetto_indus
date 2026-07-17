#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Server locale per la dashboard "Monitoraggio Attrezzature & Manutenzione Predittiva".

Cosa fa:
  - Serve la pagina monitoraggio_attrezzature.html (e i suoi asset: logo, schemi
    elettrici) dalla stessa cartella in cui si trova questo script.
  - Espone una piccola API JSON che legge/scrive dati_store.json: questo file
    e' l'unica fonte dati "viva" della dashboard, e viene aggiornato ad ogni
    modifica fatta dall'interfaccia (nuovo intervento, responsabile, email in
    copia, prelievo, nuova attrezzatura). Cosi' tutto quello che inserisci
    resta salvato anche se chiudi il browser o riavvii il PC.
  - dati_fittizi_attrezzature.xlsx NON viene piu' letto/scritto automaticamente
    da questo server: resta un file di riferimento/consultazione. Se in futuro
    serve un "reimporta da Excel" si puo' aggiungere come funzione separata.

USO:
    python server.py
    (si apre automaticamente il browser su http://localhost:8000/ )

    Per fermarlo: Ctrl+C nel terminale.
"""

import json
import re
import socket
import threading
import time
import webbrowser
from datetime import datetime
from email.parser import BytesParser
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATI_PATH = BASE_DIR / "dati_store.json"
DOC_PROCEDURE_DIR = BASE_DIR / "documenti" / "procedure"
DOC_CODICI_DIR = BASE_DIR / "documenti" / "codici_prodotto"
SCHEMI_DIR = BASE_DIR / "schemi_elettrici"
RIP_TOOL_XLSX = BASE_DIR / "riparazionitool.xlsx"
RIP_SCHEDE_XLSX = BASE_DIR / "riparazionischede.xlsx"
NAVY = "182C4C"
PORT = 8000

_lock = threading.Lock()
_data = None  # cache in memoria, sincronizzata su disco ad ogni modifica

# --- Dati fittizi di partenza per il nuovo registro schede (elettroniche) ---
SCHEDE_INIZIALI = [
    {"codice": "SCH-3001", "nome": "Scheda alimentatore ausiliario 24VCC"},
    {"codice": "SCH-3002", "nome": "Scheda driver freno elettropneumatico"},
    {"codice": "SCH-3003", "nome": "Scheda interfaccia display cabina"},
    {"codice": "SCH-3004", "nome": "Scheda encoder velocità assile"},
    {"codice": "SCH-3005", "nome": "Scheda gateway comunicazione MVB/CAN"},
    {"codice": "SCH-3006", "nome": "Scheda controllo porte automatiche"},
    {"codice": "SCH-3007", "nome": "Scheda monitoraggio temperatura motore trazione"},
    {"codice": "SCH-3008", "nome": "Scheda I/O digitale cabina"},
    {"codice": "SCH-3009", "nome": "Scheda amplificatore segnale antenna GPS"},
    {"codice": "SCH-3010", "nome": "Scheda regolatore climatizzazione"},
]

RIPARAZIONI_SCHEDE_INIZIALI = [
    {"codice_scheda": "SCH-3001", "codice_attrezzatura": "S-2214", "data": "2025-09-12", "motivo_guasto": "Componente bruciato per sovratensione", "componente_sostituito": "Regolatore di tensione LM7805", "operatore": "Luca Marino"},
    {"codice_scheda": "SCH-3001", "codice_attrezzatura": "S-2214", "data": "2026-02-03", "motivo_guasto": "Condensatore elettrolitico esaurito", "componente_sostituito": "Condensatore elettrolitico 470uF", "operatore": "Federica Galli"},
    {"codice_scheda": "SCH-3002", "codice_attrezzatura": "TLS-102", "data": "2025-08-20", "motivo_guasto": "Corto circuito su alimentazione", "componente_sostituito": "Fusibile di linea", "operatore": "Simone Ricci"},
    {"codice_scheda": "SCH-3003", "codice_attrezzatura": "DTR-014", "data": "2025-11-05", "motivo_guasto": "Saldatura a freddo su connettore", "componente_sostituito": "Connettore display", "operatore": "Elisa Bruno"},
    {"codice_scheda": "SCH-3004", "codice_attrezzatura": "TLS-118", "data": "2025-10-14", "motivo_guasto": "Interferenza elettromagnetica su sensore", "componente_sostituito": "Sensore encoder ottico", "operatore": "Luca Marino"},
    {"codice_scheda": "SCH-3004", "codice_attrezzatura": "TLS-118", "data": "2026-01-22", "motivo_guasto": "Rottura meccanica connettore", "componente_sostituito": "Connettore encoder", "operatore": "Simone Ricci"},
    {"codice_scheda": "SCH-3004", "codice_attrezzatura": "TLS-118", "data": "2026-04-30", "motivo_guasto": "Interferenza elettromagnetica su sensore", "componente_sostituito": "Sensore encoder ottico", "operatore": "Federica Galli"},
    {"codice_scheda": "SCH-3005", "codice_attrezzatura": "TLS-133", "data": "2025-09-28", "motivo_guasto": "Guasto su circuito di comunicazione", "componente_sostituito": "Transceiver CAN", "operatore": "Elisa Bruno"},
    {"codice_scheda": "SCH-3005", "codice_attrezzatura": "TLS-133", "data": "2026-03-11", "motivo_guasto": "Guasto su circuito di comunicazione", "componente_sostituito": "Transceiver CAN", "operatore": "Luca Marino"},
    {"codice_scheda": "SCH-3006", "codice_attrezzatura": "S-2255", "data": "2025-12-02", "motivo_guasto": "Rottura meccanica connettore", "componente_sostituito": "Relè di potenza", "operatore": "Simone Ricci"},
    {"codice_scheda": "SCH-3007", "codice_attrezzatura": "DTR-041", "data": "2025-08-09", "motivo_guasto": "Umidità infiltrata nel case", "componente_sostituito": "Resistore di pull-up", "operatore": "Federica Galli"},
    {"codice_scheda": "SCH-3008", "codice_attrezzatura": "S-2214", "data": "2026-01-17", "motivo_guasto": "Corto circuito su alimentazione", "componente_sostituito": "Diodo di protezione", "operatore": "Elisa Bruno"},
    {"codice_scheda": "SCH-3009", "codice_attrezzatura": "TLS-147", "data": "2025-11-27", "motivo_guasto": "Rottura meccanica connettore", "componente_sostituito": "Connettore antenna GPS", "operatore": "Luca Marino"},
    {"codice_scheda": "SCH-3010", "codice_attrezzatura": "S-2201", "data": "2026-02-19", "motivo_guasto": "Componente bruciato per sovratensione", "componente_sostituito": "Condensatore elettrolitico 470uF", "operatore": "Simone Ricci"},
    {"codice_scheda": "SCH-3010", "codice_attrezzatura": "S-2201", "data": "2026-05-06", "motivo_guasto": "Componente bruciato per sovratensione", "componente_sostituito": "Condensatore elettrolitico 470uF", "operatore": "Federica Galli"},
]


def _migra_riparazioni_guasto_da_interventi():
    """Le vecchie voci 'Riparazione guasto' dentro interventi diventano
    riparazioni_tool, cosi' lo storico esistente (usato per MTBF/tasso di
    guasto) non va perso quando si passa al nuovo modello dati."""
    rimasti = []
    for i in _data.get("interventi", []):
        if i.get("tipo") == "Riparazione guasto":
            _data["riparazioni_tool"].append({
                "codice_attrezzatura": i["codice"],
                "data": i["data"],
                "motivo_guasto": i.get("descrizione") or "Non specificato (dato migrato)",
                "componente_sostituito": "Non specificato (dato migrato)",
                "operatore": "",
            })
        else:
            rimasti.append(i)
    _data["interventi"] = rimasti


def carica_dati():
    global _data
    if DATI_PATH.exists():
        with open(DATI_PATH, encoding="utf-8") as f:
            _data = json.load(f)
    else:
        _data = {"generato_il": datetime.now().strftime("%Y-%m-%d"), "attrezzature": [], "interventi": []}

    _data.setdefault("documenti_procedure", {})
    _data.setdefault("documenti_codici_prodotto", {})

    prima_del_bootstrap_riparazioni = "riparazioni_tool" not in _data
    _data.setdefault("schede", [])
    _data.setdefault("riparazioni_tool", [])
    _data.setdefault("riparazioni_schede", [])

    if prima_del_bootstrap_riparazioni:
        _migra_riparazioni_guasto_da_interventi()
    if not _data["schede"]:
        _data["schede"] = [dict(s) for s in SCHEDE_INIZIALI]
    if not _data["riparazioni_schede"]:
        _data["riparazioni_schede"] = [dict(r) for r in RIPARAZIONI_SCHEDE_INIZIALI]

    for d in (DOC_PROCEDURE_DIR, DOC_CODICI_DIR, SCHEMI_DIR):
        d.mkdir(parents=True, exist_ok=True)

    # Bootstrap: registra automaticamente i PDF procedura gia' presenti sul disco
    # (es. quelli generati per demo) se non risultano ancora nel registro documenti.
    for pdf in sorted(DOC_PROCEDURE_DIR.glob("*.pdf")):
        codice = pdf.stem
        if codice not in _data["documenti_procedure"]:
            rel = pdf.relative_to(BASE_DIR).as_posix()
            _data["documenti_procedure"][codice] = {"nome": pdf.name, "url": rel}

    salva_dati()
    rigenera_xlsx_riparazioni()
    return _data


def sanitizza_nome_file(nome):
    nome = Path(nome or "file").name  # rimuove eventuali percorsi
    nome = re.sub(r"[^A-Za-z0-9._-]+", "_", nome)
    return nome or "file"


def parse_multipart(content_type, body):
    """Analizza un body multipart/form-data usando il modulo email della
    standard library (evita la dipendenza dal modulo 'cgi', deprecato/rimosso
    nelle versioni piu' recenti di Python)."""
    header = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8")
    msg = BytesParser().parsebytes(header + body)
    campi = {}
    if msg.is_multipart():
        for part in msg.get_payload():
            disposition = part.get("Content-Disposition", "")
            if not disposition:
                continue
            name = part.get_param("name", header="Content-Disposition")
            filename = part.get_filename()
            payload = part.get_payload(decode=True)
            if filename:
                campi[name] = {"filename": filename, "data": payload}
            else:
                campi[name] = {"filename": None, "data": (payload or b"").decode("utf-8", "replace")}
    return campi


def salva_dati():
    """Scrittura atomica: file temporaneo poi rinomina, per evitare file corrotti
    in caso di interruzione a meta' scrittura."""
    tmp = DATI_PATH.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(_data, f, ensure_ascii=False, indent=2)
    tmp.replace(DATI_PATH)


def trova_attrezzatura(codice):
    for a in _data["attrezzature"]:
        if a["codice"] == codice:
            return a
    return None


def trova_scheda(codice):
    for s in _data.get("schede", []):
        if s["codice"] == codice:
            return s
    return None


def _intestazione_xlsx(ws, headers):
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)


def _scrivi_riparazionitool_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riparazioni Tool"
    _intestazione_xlsx(ws, ["Data", "Codice Attrezzatura", "Nome Attrezzatura", "Motivo Guasto", "Componente Sostituito", "Operatore"])
    nomi = {a["codice"]: a["nome"] for a in _data["attrezzature"]}
    for r in sorted(_data.get("riparazioni_tool", []), key=lambda x: x["data"]):
        ws.append([r["data"], r["codice_attrezzatura"], nomi.get(r["codice_attrezzatura"], ""),
                   r["motivo_guasto"], r["componente_sostituito"], r.get("operatore", "")])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="Segoe UI")
    for i, w in enumerate([14, 18, 38, 34, 30, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    tmp = RIP_TOOL_XLSX.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    tmp.replace(RIP_TOOL_XLSX)


def _scrivi_riparazionischede_xlsx():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Schede"
    _intestazione_xlsx(ws1, ["Codice Scheda", "Nome Scheda"])
    for s in sorted(_data.get("schede", []), key=lambda x: x["codice"]):
        ws1.append([s["codice"], s["nome"]])
    for row in ws1.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="Segoe UI")
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 44

    ws2 = wb.create_sheet("Riparazioni")
    _intestazione_xlsx(ws2, ["Data", "Codice Scheda", "Nome Scheda", "Attrezzatura Utilizzata", "Motivo Guasto", "Componente Sostituito", "Operatore"])
    nomi_schede = {s["codice"]: s["nome"] for s in _data.get("schede", [])}
    for r in sorted(_data.get("riparazioni_schede", []), key=lambda x: x["data"]):
        ws2.append([r["data"], r["codice_scheda"], nomi_schede.get(r["codice_scheda"], ""),
                    r["codice_attrezzatura"], r["motivo_guasto"], r["componente_sostituito"], r.get("operatore", "")])
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name="Segoe UI")
    for i, w in enumerate([14, 14, 38, 20, 34, 30, 18], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    tmp = RIP_SCHEDE_XLSX.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    tmp.replace(RIP_SCHEDE_XLSX)


def rigenera_xlsx_riparazioni():
    """Rigenera le viste Excel di sola consultazione per le riparazioni.
    La fonte dati viva resta sempre dati_store.json; se uno dei due file e'
    aperto in Excel in questo momento l'aggiornamento viene semplicemente
    saltato con un avviso, senza bloccare il server e senza corrompere nulla."""
    try:
        _scrivi_riparazionitool_xlsx()
    except PermissionError:
        print(f"Attenzione: {RIP_TOOL_XLSX.name} e' aperto altrove, aggiornamento saltato.")
    except Exception as e:
        print(f"Errore aggiornando {RIP_TOOL_XLSX.name}: {e}")

    try:
        _scrivi_riparazionischede_xlsx()
    except PermissionError:
        print(f"Attenzione: {RIP_SCHEDE_XLSX.name} e' aperto altrove, aggiornamento saltato.")
    except Exception as e:
        print(f"Errore aggiornando {RIP_SCHEDE_XLSX.name}: {e}")


def costruisci_log_mensile(ore_attuali):
    log = []
    cum = 0
    oggi = datetime.now()
    for i in range(11, -1, -1):
        anno = oggi.year
        mese = oggi.month - i
        while mese <= 0:
            mese += 12
            anno -= 1
        inc = round(ore_attuali / 12)
        cum += inc
        log.append({"mese": f"{anno}-{mese:02d}", "ore_incrementali": inc, "ore_cumulative": cum})
    if log:
        log[-1]["ore_cumulative"] = ore_attuali
    return log


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def log_message(self, fmt, *args):
        print(f"[{datetime.now():%H:%M:%S}] {args[0]} {args[1]}" if len(args) >= 2 else fmt % args)

    def _json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        # Consentito esplicitamente da qualunque origine: la dashboard viene aperta
        # anche da smartphone tramite l'IP di rete del PC, non solo da localhost.
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _body_json(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    # ---------------- GET ----------------
    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/api/data":
            with _lock:
                self._json(_data)
            return
        if path == "/":
            self.path = "/monitoraggio_attrezzature.html"
        return super().do_GET()

    # ---------------- POST ----------------
    def do_POST(self):
        path = urlparse(self.path).path
        content_type = self.headers.get("Content-Type", "")
        try:
            with _lock:
                if content_type.startswith("multipart/form-data"):
                    length = int(self.headers.get("Content-Length", 0))
                    raw = self.rfile.read(length) if length else b""
                    campi = parse_multipart(content_type, raw)
                    if path == "/api/upload/procedura":
                        self._api_upload_procedura(campi)
                    elif path == "/api/upload/schema":
                        self._api_upload_schema(campi)
                    elif path == "/api/upload/codice_prodotto":
                        self._api_upload_codice_prodotto(campi)
                    else:
                        self._json({"ok": False, "errore": "Endpoint sconosciuto"}, status=404)
                        return
                else:
                    body = self._body_json()
                    if path == "/api/intervento":
                        self._api_intervento(body)
                    elif path == "/api/responsabile":
                        self._api_responsabile(body)
                    elif path == "/api/cc/add":
                        self._api_cc_add(body)
                    elif path == "/api/cc/remove":
                        self._api_cc_remove(body)
                    elif path == "/api/prelievo":
                        self._api_prelievo(body)
                    elif path == "/api/attrezzatura":
                        self._api_nuova_attrezzatura(body)
                    elif path == "/api/schema":
                        self._api_schema(body)
                    elif path == "/api/riparazione-tool":
                        self._api_riparazione_tool(body)
                    elif path == "/api/riparazione-scheda":
                        self._api_riparazione_scheda(body)
                    else:
                        self._json({"ok": False, "errore": "Endpoint sconosciuto"}, status=404)
                        return
                salva_dati()
                self._json({"ok": True, "data": _data})
        except ValueError as e:
            self._json({"ok": False, "errore": str(e)}, status=400)
        except Exception as e:
            self._json({"ok": False, "errore": f"Errore interno: {e}"}, status=500)

    # ---------------- logica endpoint ----------------
    def _api_intervento(self, body):
        codice = body.get("codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        data_i = body.get("data") or datetime.now().strftime("%Y-%m-%d")
        tipo = body.get("tipo", "Manutenzione ordinaria")
        esito = body.get("esito", "Positivo")
        descrizione = (body.get("descrizione") or "").strip() or "—"

        # Se e' una taratura annuale con esito positivo, la scadenza viene ricalcolata
        # (+1 anno dalla data dell'intervento). Per non perdere la cronistoria in fase
        # di audit qualita', la vecchia e la nuova scadenza vengono annotate esplicitamente
        # dentro la descrizione dell'intervento salvato nello storico, cosi' resta scritto
        # nero su bianco cosa e' cambiato e quando, anche a distanza di anni.
        nota_audit = ""
        if tipo == "Taratura annuale" and esito == "Positivo":
            vecchia_scadenza = raw.get("scadenza_revisione")
            d = datetime.strptime(data_i, "%Y-%m-%d")
            nuova_scadenza = d.replace(year=d.year + 1).strftime("%Y-%m-%d")
            raw["ultimo_controllo"] = data_i
            raw["scadenza_revisione"] = nuova_scadenza
            if vecchia_scadenza and vecchia_scadenza != nuova_scadenza:
                nota_audit = f" [Scadenza aggiornata: {vecchia_scadenza} → {nuova_scadenza}]"
            else:
                nota_audit = f" [Nuova scadenza: {nuova_scadenza}]"

        _data["interventi"].append({
            "codice": codice, "data": data_i, "tipo": tipo,
            "descrizione": descrizione + nota_audit, "esito": esito,
        })

    def _api_responsabile(self, body):
        codice = body.get("codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        raw["responsabile_nome"] = (body.get("nome") or "").strip() or "Da assegnare"
        raw["responsabile_email"] = (body.get("email") or "").strip()

    def _api_cc_add(self, body):
        codice = body.get("codice")
        email = (body.get("email") or "").strip()
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        if not email:
            raise ValueError("Email vuota")
        raw.setdefault("email_cc", []).append(email)

    def _api_cc_remove(self, body):
        codice = body.get("codice")
        idx = body.get("idx")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        cc = raw.get("email_cc", [])
        if not isinstance(idx, int) or not (0 <= idx < len(cc)):
            raise ValueError("Indice CC non valido")
        cc.pop(idx)

    def _api_prelievo(self, body):
        codice = body.get("codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        raw.setdefault("prelievi", []).append({
            "data": body.get("data") or datetime.now().strftime("%Y-%m-%d"),
            "richiesto_da": (body.get("richiesto_da") or "").strip(),
            "bloccante": bool(body.get("bloccante")),
            "nota": (body.get("nota") or "").strip(),
        })

    def _api_nuova_attrezzatura(self, body):
        codice = (body.get("codice") or "").strip()
        nome = (body.get("nome") or "").strip()
        scadenza = (body.get("scadenza_revisione") or "").strip()
        soglia = body.get("soglia_ore")
        if not codice or not nome or not scadenza or not soglia:
            raise ValueError("Campi obbligatori mancanti (codice, nome, scadenza, soglia ore)")
        if trova_attrezzatura(codice):
            raise ValueError(f"Il codice '{codice}' e' gia' presente")

        ore_attuali = body.get("ore_uso_attuali") or 0
        _data["attrezzature"].append({
            "codice": codice,
            "nome": nome,
            "ubicazione": (body.get("ubicazione") or "").strip() or "Da definire",
            "ultimo_controllo": (body.get("ultimo_controllo") or "").strip() or datetime.now().strftime("%Y-%m-%d"),
            "scadenza_revisione": scadenza,
            "soglia_ore": float(soglia),
            "ore_uso_attuali": float(ore_attuali),
            "log_ore_mensile": costruisci_log_mensile(float(ore_attuali)),
            "guasti_12_mesi": int(body.get("guasti_12_mesi") or 0),
            "codici_prodotto": body.get("codici_prodotto") or [],
            "procedure_collaudo": body.get("procedure_collaudo") or [],
            "responsabile_nome": (body.get("responsabile_nome") or "").strip() or "Da assegnare",
            "responsabile_email": (body.get("responsabile_email") or "").strip(),
            "email_cc": [],
            "schemi_elettrici": [],
            "prelievi": [],
        })

    def _api_schema(self, body):
        """Registra un riferimento a uno schema elettrico gia' salvato manualmente
        nella cartella schemi_elettrici/ (il caricamento file vero e proprio dal
        browser resta locale alla sessione, per limiti di sicurezza del browser)."""
        codice = body.get("codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        nome = (body.get("nome") or "").strip()
        url = (body.get("url") or "").strip()
        if not nome or not url:
            raise ValueError("nome/url mancanti")
        raw.setdefault("schemi_elettrici", []).append({"nome": nome, "url": url})

    def _api_riparazione_tool(self, body):
        """Registra una riparazione/guasto del TOOL stesso (non di una scheda).
        Questa e' la fonte del tasso di guasto/MTBF dell'attrezzatura."""
        codice = body.get("codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        _data.setdefault("riparazioni_tool", []).append({
            "codice_attrezzatura": codice,
            "data": body.get("data") or datetime.now().strftime("%Y-%m-%d"),
            "motivo_guasto": (body.get("motivo_guasto") or "").strip() or "Non specificato",
            "componente_sostituito": (body.get("componente_sostituito") or "").strip() or "Non specificato",
            "operatore": (body.get("operatore") or "").strip(),
        })
        rigenera_xlsx_riparazioni()

    def _api_riparazione_scheda(self, body):
        """Registra la riparazione di una SCHEDA (entita' separata dalle
        attrezzature) effettuata usando un tool: incrementa il fattore di
        guasto della scheda, non quello del tool usato per ripararla.
        Se il codice scheda non esiste ancora, viene creato al volo se e'
        stato fornito anche il nome."""
        codice_scheda = (body.get("codice_scheda") or "").strip()
        codice_attrezzatura = body.get("codice_attrezzatura")
        if not codice_scheda:
            raise ValueError("codice_scheda mancante")
        if not trova_attrezzatura(codice_attrezzatura):
            raise ValueError(f"Attrezzatura '{codice_attrezzatura}' non trovata")

        if not trova_scheda(codice_scheda):
            nome_scheda = (body.get("nome_scheda") or "").strip()
            if not nome_scheda:
                raise ValueError(f"La scheda '{codice_scheda}' non esiste: indica anche il nome per crearla")
            _data.setdefault("schede", []).append({"codice": codice_scheda, "nome": nome_scheda})

        _data.setdefault("riparazioni_schede", []).append({
            "codice_scheda": codice_scheda,
            "codice_attrezzatura": codice_attrezzatura,
            "data": body.get("data") or datetime.now().strftime("%Y-%m-%d"),
            "motivo_guasto": (body.get("motivo_guasto") or "").strip() or "Non specificato",
            "componente_sostituito": (body.get("componente_sostituito") or "").strip() or "Non specificato",
            "operatore": (body.get("operatore") or "").strip(),
        })
        rigenera_xlsx_riparazioni()

    # ---------------- upload reale di file (multipart/form-data) ----------------
    def _campo_testo(self, campi, nome):
        c = campi.get(nome)
        if not c or c.get("filename") is not None:
            return ""
        return (c.get("data") or "").strip()

    def _campo_file(self, campi, nome="file"):
        c = campi.get(nome)
        if not c or not c.get("filename"):
            raise ValueError("Nessun file ricevuto")
        return c["filename"], c["data"] or b""

    def _api_upload_procedura(self, campi):
        codice_proc = self._campo_testo(campi, "codice_procedura")
        if not codice_proc:
            raise ValueError("codice_procedura mancante")
        filename, contenuto = self._campo_file(campi)
        nome_salvato = f"{sanitizza_nome_file(codice_proc)}__{sanitizza_nome_file(filename)}"
        destinazione = DOC_PROCEDURE_DIR / nome_salvato
        destinazione.write_bytes(contenuto)
        rel = destinazione.relative_to(BASE_DIR).as_posix()
        _data["documenti_procedure"][codice_proc] = {"nome": filename, "url": rel}

    def _api_upload_codice_prodotto(self, campi):
        codice_prod = self._campo_testo(campi, "codice_prodotto")
        if not codice_prod:
            raise ValueError("codice_prodotto mancante")
        filename, contenuto = self._campo_file(campi)
        lista = _data["documenti_codici_prodotto"].setdefault(codice_prod, [])
        nome_salvato = f"{sanitizza_nome_file(codice_prod)}__{len(lista)}__{sanitizza_nome_file(filename)}"
        destinazione = DOC_CODICI_DIR / nome_salvato
        destinazione.write_bytes(contenuto)
        rel = destinazione.relative_to(BASE_DIR).as_posix()
        lista.append({"nome": filename, "url": rel})

    def _api_upload_schema(self, campi):
        codice = self._campo_testo(campi, "codice")
        raw = trova_attrezzatura(codice)
        if not raw:
            raise ValueError(f"Codice '{codice}' non trovato")
        filename, contenuto = self._campo_file(campi)
        lista = raw.setdefault("schemi_elettrici", [])
        nome_salvato = f"{sanitizza_nome_file(codice)}__{len(lista)}__{sanitizza_nome_file(filename)}"
        destinazione = SCHEMI_DIR / nome_salvato
        destinazione.write_bytes(contenuto)
        rel = destinazione.relative_to(BASE_DIR).as_posix()
        lista.append({"nome": filename, "url": rel})


def ottieni_ip_locali():
    """Trova gli indirizzi IPv4 di questo PC sulla rete locale (Wi-Fi/LAN), per
    poterli comunicare a chi vuole aprire la dashboard da smartphone/tablet
    collegato alla stessa rete. Se il PC ha piu' schede di rete (es. VPN, o
    adattatori virtuali creati da Docker/Hyper-V/VMware) puo' comparire piu'
    di un indirizzo: in quel caso va provato quello che assomiglia all'IP del
    router di casa (tipicamente 192.168.x.x)."""
    trovati = []

    # Trucco standard: "connettersi" (senza inviare nulla) a un indirizzo
    # esterno fa scegliere al sistema operativo l'interfaccia di rete che
    # userebbe per uscire su internet - di solito e' quella giusta.
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        trovati.append(s.getsockname()[0])
    except Exception:
        pass
    finally:
        s.close()

    # In aggiunta, elenchiamo anche tutti gli altri indirizzi IPv4 associati
    # al nome host: utile come alternativa se il primo non fosse quello giusto.
    try:
        for ip in socket.gethostbyname_ex(socket.gethostname())[2]:
            if ip not in trovati and not ip.startswith("127."):
                trovati.append(ip)
    except Exception:
        pass

    return trovati


def main():
    carica_dati()
    print(f"Dati caricati da: {DATI_PATH}")
    print(f"Attrezzature in memoria: {len(_data['attrezzature'])}")

    # 0.0.0.0 = in ascolto su tutte le interfacce di rete, non solo su questo PC:
    # cosi' la dashboard e' raggiungibile anche da smartphone/tablet sulla stessa
    # rete Wi-Fi, non solo aprendo il browser sul PC dove gira il server.
    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    url = f"http://localhost:{PORT}/"
    ip_locali = ottieni_ip_locali()

    print("Server avviato.")
    print(f"  Su questo PC:                {url}")
    if ip_locali:
        print("  Da smartphone/tablet (stessa rete Wi-Fi), prova questo indirizzo:")
        print(f"    http://{ip_locali[0]}:{PORT}/")
        if len(ip_locali) > 1:
            print("  Se non funziona, il PC ha piu' schede di rete: prova anche:")
            for ip in ip_locali[1:]:
                print(f"    http://{ip}:{PORT}/")
    else:
        print("  Da smartphone/tablet: non e' stato possibile rilevare l'IP di rete automaticamente.")
        print("  Cerca l'indirizzo IPv4 del PC (Impostazioni > Rete, oppure 'ipconfig' nel prompt dei comandi,")
        print(f"   cerca la riga 'Indirizzo IPv4' della scheda Wi-Fi) e apri http://<QUELL'IP>:{PORT}/ dal telefono.")
    print()
    print("  Se lo smartphone non riesce a collegarsi (pagina bianca o 'impossibile contattare")
    print("  il server'), la causa piu' comune e' il Firewall di Windows che blocca le connessioni")
    print("  in ingresso su questa porta. Fai doppio click su 'consenti_firewall.bat' in questa")
    print("  cartella, scegli 'Si'/'Esegui come amministratore' se richiesto, poi riprova dal telefono.")
    print("  (Verifica anche che telefono e PC siano collegati alla stessa rete Wi-Fi.)")
    print("Premi Ctrl+C per fermarlo.")

    threading.Timer(0.6, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer arrestato.")
        server.shutdown()


if __name__ == "__main__":
    main()
