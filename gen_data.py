import random
from datetime import date, timedelta
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

random.seed(42)

OUT_DIR = "/sessions/clever-wonderful-shannon/mnt/outputs"
IMG_DIR = f"{OUT_DIR}/attrezzature_foto"

# -----------------------------
# 1. Dati anagrafica attrezzature (fittizi)
# -----------------------------
attrezzature = [
    {"codice": "DTR-014", "nome": "Multimetro digitale DMM",        "colore": (59,130,246)},
    {"codice": "DTR-027", "nome": "Multimetro digitale DMM",        "colore": (59,130,246)},
    {"codice": "TLS-102", "nome": "Banco di prova elettrico",       "colore": (16,185,129)},
    {"codice": "TLS-118", "nome": "Banco di prova pneumatico",      "colore": (5,150,105)},
    {"codice": "S-2201",  "nome": "Chiave dinamometrica",           "colore": (245,158,11)},
    {"codice": "S-2214",  "nome": "Calibro digitale",               "colore": (234,88,12)},
    {"codice": "TLS-133", "nome": "Cella di carico",                "colore": (139,92,246)},
    {"codice": "DTR-041", "nome": "Oscilloscopio portatile",        "colore": (37,99,235)},
    {"codice": "S-2255",  "nome": "Comparatore centesimale",        "colore": (217,119,6)},
    {"codice": "TLS-147", "nome": "Termocamera",                    "colore": (220,38,38)},
]

ubicazioni_interne = ["Interna - Reparto Collaudo", "Interna - Reparto Industrializzazione", "Interna - Magazzino Attrezzature"]
fornitori = ["Presso fornitore - MetroCal Srl", "Presso fornitore - TarLab Taratura", "Presso fornitore - Precisa Strumenti"]

def random_date(start_year=2024, end_year=2025):
    start = date(start_year, 1, 1)
    end = date(end_year, 12, 31)
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))

for a in attrezzature:
    ultimo_controllo = random_date(2024, 2025)
    scadenza = ultimo_controllo.replace(year=ultimo_controllo.year + 1)
    a["ultimo_controllo"] = ultimo_controllo
    a["scadenza_revisione"] = scadenza
    a["ubicazione"] = random.choice(ubicazioni_interne + fornitori)
    a["foto_file"] = f"{a['codice']}.png"

# -----------------------------
# 2. Genera immagini placeholder (fittizie)
# -----------------------------
def make_placeholder(codice, nome, colore, path):
    W, H = 240, 180
    img = Image.new("RGB", (W, H), color=(245, 245, 247))
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, W-1, H-1], outline=(200,200,200), width=2)
    d.rectangle([20, 20, W-20, H-60], fill=colore)
    try:
        font = ImageFont.load_default()
    except Exception:
        font = None
    d.text((30, H-50), nome[:28], fill=(30,30,30), font=font)
    d.text((30, H-32), codice, fill=(80,80,80), font=font)
    d.text((30, 30), "FOTO FITTIZIA", fill=(255,255,255), font=font)
    img.save(path)

for a in attrezzature:
    make_placeholder(a["codice"], a["nome"], a["colore"], f"{IMG_DIR}/{a['foto_file']}")

# -----------------------------
# 3. Codici prodotto su cui sono utilizzate (many-to-many)
# -----------------------------
codici_prodotto_pool = [f"COD-{n}" for n in range(1000, 1060, 4)]

attrezzatura_codici = []
for a in attrezzature:
    n_codici = random.randint(1, 4)
    codici = random.sample(codici_prodotto_pool, n_codici)
    for c in codici:
        attrezzatura_codici.append({"codice_attrezzatura": a["codice"], "codice_prodotto": c})

# -----------------------------
# 4. Procedure di collaudo che richiamano l'attrezzatura (many-to-many)
# -----------------------------
procedure_pool = [f"PC-{n:04d}" for n in range(100, 160, 3)]

attrezzatura_procedure = []
for a in attrezzature:
    n_proc = random.randint(1, 3)
    procs = random.sample(procedure_pool, n_proc)
    for p in procs:
        attrezzatura_procedure.append({"codice_attrezzatura": a["codice"], "procedura_collaudo": p})

# -----------------------------
# 5. Interventi effettuati (storico)
# -----------------------------
tipi_intervento = ["Taratura annuale", "Manutenzione ordinaria", "Riparazione guasto", "Verifica straordinaria"]
esiti = ["Positivo", "Positivo", "Positivo", "Negativo - da rifare"]
descr_pool = {
    "Taratura annuale": "Taratura eseguita da fornitore esterno, certificato archiviato.",
    "Manutenzione ordinaria": "Pulizia e verifica funzionale di routine.",
    "Riparazione guasto": "Sostituito componente danneggiato, ripristinata funzionalita'.",
    "Verifica straordinaria": "Controllo dopo segnalazione operatore.",
}

interventi = []
for a in attrezzature:
    n_interventi = random.randint(2, 4)
    for _ in range(n_interventi):
        tipo = random.choice(tipi_intervento)
        interventi.append({
            "codice_attrezzatura": a["codice"],
            "data_intervento": random_date(2023, 2025),
            "tipo_intervento": tipo,
            "descrizione": descr_pool[tipo],
            "esito": random.choice(esiti),
        })
interventi.sort(key=lambda r: (r["codice_attrezzatura"], r["data_intervento"]))

# -----------------------------
# 6. Costruzione workbook
# -----------------------------
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER

def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 0: Legenda ---
ws0 = wb.active
ws0.title = "Legenda"
ws0["A1"] = "Dataset fittizio - Attrezzature di collaudo/misura"
ws0["A1"].font = Font(name="Arial", bold=True, size=13)
ws0["A3"] = "Tutti i dati in questo file sono INVENTATI a scopo dimostrativo (nessun dato aziendale reale)."
ws0["A3"].font = Font(name="Arial", italic=True, size=10)
righe = [
    ("Foglio", "Contenuto"),
    ("Attrezzature", "Anagrafica: codice, nome, foto, ubicazione, ultimo controllo, scadenza revisione"),
    ("Attrezzatura-Codici Prodotto", "Relazione N:N - su quali codici prodotto e' usata ogni attrezzatura"),
    ("Attrezzatura-Procedure Collaudo", "Relazione N:N - in quali procedure di collaudo e' richiamata ogni attrezzatura"),
    ("Interventi", "Storico interventi (taratura, manutenzione, riparazione) per attrezzatura"),
]
ws0["A5"] = "Struttura del file"
ws0["A5"].font = Font(name="Arial", bold=True, size=11)
for i, (a, b) in enumerate(righe, start=6):
    ws0.cell(row=i, column=1, value=a).font = Font(name="Arial", bold=(i==6), size=10)
    ws0.cell(row=i, column=2, value=b).font = Font(name="Arial", size=10)
autofit(ws0, [30, 75])

# --- Sheet 1: Attrezzature ---
ws1 = wb.create_sheet("Attrezzature")
headers1 = ["Codice Attrezzatura", "Nome Attrezzatura", "Foto", "Ubicazione", "Ultimo Controllo", "Scadenza Revisione"]
ws1.append(headers1)
style_header(ws1, len(headers1))
ws1.row_dimensions[1].height = 30

ROW_H = 100
for r, a in enumerate(attrezzature, start=2):
    ws1.cell(row=r, column=1, value=a["codice"]).font = BODY_FONT
    ws1.cell(row=r, column=2, value=a["nome"]).font = BODY_FONT
    ws1.cell(row=r, column=4, value=a["ubicazione"]).font = BODY_FONT
    c5 = ws1.cell(row=r, column=5, value=a["ultimo_controllo"]); c5.font = BODY_FONT; c5.number_format = "DD/MM/YYYY"
    c6 = ws1.cell(row=r, column=6, value=a["scadenza_revisione"]); c6.font = BODY_FONT; c6.number_format = "DD/MM/YYYY"
    for col in range(1, len(headers1)+1):
        ws1.cell(row=r, column=col).border = BORDER
    ws1.row_dimensions[r].height = ROW_H
    img = XLImage(f"{IMG_DIR}/{a['foto_file']}")
    img.width = 130
    img.height = 95
    ws1.add_image(img, f"C{r}")

autofit(ws1, [18, 30, 20, 32, 16, 16])

# --- Sheet 2: Attrezzatura-Codici Prodotto ---
ws2 = wb.create_sheet("Attrezzatura-Codici Prodotto")
headers2 = ["Codice Attrezzatura", "Codice Prodotto"]
ws2.append(headers2)
style_header(ws2, len(headers2))
for r, row in enumerate(attrezzatura_codici, start=2):
    ws2.cell(row=r, column=1, value=row["codice_attrezzatura"]).font = BODY_FONT
    ws2.cell(row=r, column=2, value=row["codice_prodotto"]).font = BODY_FONT
    for col in range(1, 3):
        ws2.cell(row=r, column=col).border = BORDER
autofit(ws2, [20, 20])

# --- Sheet 3: Attrezzatura-Procedure Collaudo ---
ws3 = wb.create_sheet("Attrezzatura-Procedure Collaudo")
headers3 = ["Codice Attrezzatura", "Procedura di Collaudo"]
ws3.append(headers3)
style_header(ws3, len(headers3))
for r, row in enumerate(attrezzatura_procedure, start=2):
    ws3.cell(row=r, column=1, value=row["codice_attrezzatura"]).font = BODY_FONT
    ws3.cell(row=r, column=2, value=row["procedura_collaudo"]).font = BODY_FONT
    for col in range(1, 3):
        ws3.cell(row=r, column=col).border = BORDER
autofit(ws3, [20, 24])

# --- Sheet 4: Interventi ---
ws4 = wb.create_sheet("Interventi")
headers4 = ["Codice Attrezzatura", "Data Intervento", "Tipo Intervento", "Descrizione", "Esito"]
ws4.append(headers4)
style_header(ws4, len(headers4))
for r, row in enumerate(interventi, start=2):
    ws4.cell(row=r, column=1, value=row["codice_attrezzatura"]).font = BODY_FONT
    c2 = ws4.cell(row=r, column=2, value=row["data_intervento"]); c2.font = BODY_FONT; c2.number_format = "DD/MM/YYYY"
    ws4.cell(row=r, column=3, value=row["tipo_intervento"]).font = BODY_FONT
    ws4.cell(row=r, column=4, value=row["descrizione"]).font = BODY_FONT
    esito_cell = ws4.cell(row=r, column=5, value=row["esito"])
    esito_cell.font = BODY_FONT
    if "Negativo" in row["esito"]:
        esito_cell.font = Font(name="Arial", size=10, color="B91C1C", bold=True)
    for col in range(1, 6):
        ws4.cell(row=r, column=col).border = BORDER
autofit(ws4, [20, 16, 22, 45, 20])

wb.save(f"{OUT_DIR}/dati_fittizi_attrezzature.xlsx")
print("OK - file salvato")
print("Attrezzature:", len(attrezzature))
print("Righe Attrezzatura-Codici:", len(attrezzatura_codici))
print("Righe Attrezzatura-Procedure:", len(attrezzatura_procedure))
print("Righe Interventi:", len(interventi))
